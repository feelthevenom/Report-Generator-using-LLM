import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import streamlit as st
from utils.helpers import ordinal_suffix, parse_descriptions
from services.llm_service import summarize_task
from config.settings import EXCEL_STYLES

def generate_excel_timesheet(employee_info, date_range, descriptions_text, dates, progress_bar, status_text):
    """
    Generate an Excel timesheet based on the provided information
    
    Args:
        employee_info (dict): Employee information dictionary
        date_range (tuple): Tuple containing start_date and end_date
        descriptions_text (str): Text from the descriptions text area
        dates (list): List of dates within the range
        progress_bar (streamlit.delta_generator.DeltaGenerator): Streamlit progress bar
        status_text (streamlit.delta_generator.DeltaGenerator): Streamlit text for status updates
    
    Returns:
        tuple: (io.BytesIO with Excel file, list of preview data dictionaries)
    """
    start_date, end_date = date_range
    
    # Parse descriptions
    date_to_desc = parse_descriptions(descriptions_text, dates)
    
    # Check if any predefined holidays fall within the date range
    holiday_dict = {}
    for holiday in st.session_state.holidays:
        try:
            if isinstance(holiday, dict):
                holiday_date = holiday['date']
                if isinstance(holiday_date, str):
                    import datetime
                    holiday_date = datetime.datetime.strptime(holiday_date, '%Y-%m-%d').date()
                if start_date <= holiday_date <= end_date:
                    holiday_dict[holiday_date] = holiday['name']
        except (ValueError, KeyError):
            continue  # Skip invalid holiday entries
    
    # Create Excel workbook with styling
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    
    # Define styles
    header_font = Font(bold=True, size=EXCEL_STYLES['header_font_size'])
    header_fill = PatternFill(
        start_color=EXCEL_STYLES['header_bg_color'], 
        end_color=EXCEL_STYLES['header_bg_color'], 
        fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_alignment = Alignment(vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    weekend_fill = PatternFill(
        start_color=EXCEL_STYLES['weekend_bg_color'], 
        end_color=EXCEL_STYLES['weekend_bg_color'], 
        fill_type="solid"
    )
    holiday_fill = PatternFill(
        start_color=EXCEL_STYLES['holiday_bg_color'], 
        end_color=EXCEL_STYLES['holiday_bg_color'], 
        fill_type="solid"
    )
    leave_fill = PatternFill(
        start_color=EXCEL_STYLES['leave_bg_color'], 
        end_color=EXCEL_STYLES['leave_bg_color'], 
        fill_type="solid"
    )
    
    # Add title and employee information
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Timesheet: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    ws['A1'].font = Font(bold=True, size=EXCEL_STYLES['title_font_size'])
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:C2')
    ws['A2'] = f"Employee: {employee_info['name']}"
    ws['A2'].font = Font(bold=True)
    
    ws.merge_cells('D2:E2')
    ws['D2'] = f"Designation: {employee_info['designation']}"
    ws['D2'].font = Font(bold=True)
    
    ws.merge_cells('F2:G2')
    ws['F2'] = f"Department: {employee_info['department']}"
    ws['F2'].font = Font(bold=True)
    
    ws.merge_cells('A3:C3')
    ws['A3'] = f"Employee ID: {employee_info['employee_id']}"
    ws['A3'].font = Font(bold=True)
    
    # Empty row
    ws.append([])
    
    # Write headers on row 5
    headers = [
        "Date", "Day", "Status", "Start Time", "End Time",
        "Duration (hrs)", "Task Details"
    ]
    ws.append(headers)
    
    # Style headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Adjust column widths
    column_widths = EXCEL_STYLES['column_widths']
    ws.column_dimensions['A'].width = column_widths['date']
    ws.column_dimensions['B'].width = column_widths['day']
    ws.column_dimensions['C'].width = column_widths['status']
    ws.column_dimensions['D'].width = column_widths['start_time']
    ws.column_dimensions['E'].width = column_widths['end_time']
    ws.column_dimensions['F'].width = column_widths['duration']
    ws.column_dimensions['G'].width = column_widths['task']
    
    # Preview data
    preview_data = []
    
    # Fill in rows for each date
    for i, date in enumerate(dates):
        row_num = i + 6  # Starting from row 6 after headers
        
        status_text.text(f"Processing date {i+1}/{len(dates)}")
        progress_bar.progress(0.5 + (i + 1) / (len(dates) * 2))  # Second half for Excel generation
        
        description = date_to_desc.get(date, "")
        
        # Initialize row data
        status = ""
        start_time = ""
        end_time = ""
        duration = ""
        task = ""
        
        # Format date and day
        day_suffix = ordinal_suffix(date.day)
        date_str = f"{date.day}{day_suffix} {date.strftime('%b %Y')}"
        day_name = date.strftime('%a')
        
        # Determine if it's a working day
        if date in holiday_dict:
            status = f"Holiday - {holiday_dict[date]}"
            row_fill = holiday_fill
        elif description.lower().startswith("holiday"):
            status = "Holiday"
            if len(description) > 7:  # If there's more text after "holiday"
                status = description
            row_fill = holiday_fill
        elif description.lower().startswith("leave"):
            status = "Leave"
            if len(description) > 5:  # If there's more text after "leave"
                status = description
            row_fill = leave_fill
        elif description.lower().startswith("weekend") or date.weekday() >= 5:
            status = "Weekend"
            row_fill = weekend_fill
        elif description.strip() == "":
            if date.weekday() >= 5:  # Saturday (5) or Sunday (6)
                status = "Weekend"
                row_fill = weekend_fill
            else:
                status = "Regular Day"
                row_fill = None
        else:
            # Working day
            status = "Regular Day"
            start_time = st.session_state.work_hours['start']
            end_time = st.session_state.work_hours['end']
            duration = st.session_state.work_hours['duration']
            row_fill = None
            
            # Use LLM to summarize task
            if description and not description.lower().startswith(("holiday", "leave", "weekend")):
                task = summarize_task(description)
            else:
                task = description
        
        # Write row to Excel
        row_data = [date_str, day_name, status, start_time, end_time, duration, task]
        ws.append(row_data)
        
        # Add to preview data
        preview_data.append({
            "Date": date_str,
            "Day": day_name,
            "Status": status,
            "Task": task if task else "N/A"
        })
        
        # Style the row
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = normal_alignment
            if row_fill:
                cell.fill = row_fill
    
    # Add signature section
    ws.append([])
    last_row = len(dates) + 7
    
    ws.merge_cells(f'A{last_row}:C{last_row}')
    ws[f'A{last_row}'] = "Employee Signature"
    ws[f'A{last_row}'].font = Font(bold=True)
    
    ws.merge_cells(f'E{last_row}:G{last_row}')
    ws[f'E{last_row}'] = "Manager Signature"
    ws[f'E{last_row}'].font = Font(bold=True)
    
    # Save to BytesIO for download
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Clear progress indicators
    progress_bar.empty()
    status_text.empty()
    
    return output, preview_data